from openpyxl import load_workbook
import openpyxl
from User import User
import pyexcel as p
import PySimpleGUI as sg
import pathlib
from pathlib import Path


def create_user_list(file):
    us = list()
    s_num = -1
    n_num = -1
    sn_num = -1
    for i in range(1, file.max_row + 1):
        for j in range(1, file.max_column + 1):
            cell1 = file.cell(row=i, column=j)
            if cell1.value == 'Фамилия' or cell1.value == 'фамилия':
                s_num = j
            elif cell1.value == 'Имя' or cell1.value == 'имя':
                n_num = j
            elif cell1.value == 'Отчество' or cell1.value == 'отчество':
                sn_num = j
            elif j == s_num and i != 1:
                surname = cell1.value
            elif j == n_num and i != 1:
                name = cell1.value
            elif j == sn_num and i != 1:
                second_name = cell1.value
        if i != 1:
            us.append(User(surname, name, second_name))
    return us


def main():
    layout = [
        [sg.Text('Файл для выгрузки  '), sg.InputText(), sg.FileBrowse(),
         ],
        [sg.Text('Файл для сравнения'), sg.InputText(), sg.FileBrowse(),
         ],
        [sg.Text('Итоговый файл        '), sg.InputText(), sg.FileBrowse(),
         ],
        [sg.Submit(), sg.Cancel()]
    ]
    window = sg.Window('File Compare', layout)
    try:
        while True:
            event, values = window.read()
            if event in (None, 'Exit', 'Cancel'):
                break
            try:
                wb = openpyxl.Workbook()
            except Exception as err:
                print(err)
                break
            wb1 = openpyxl.load_workbook(values[0])
            wb2 = openpyxl.load_workbook(values[1])
            wb.create_sheet(title='Лист1', index=0)
            sheet = wb['Лист1']
            sheet1 = wb1['Лист1']
            sheet2 = wb2['Лист1']
            print(sheet1.max_column)
            for i in range(1, sheet1.max_column + 1):
                cell = sheet.cell(row=1, column=i)
                cell1 = sheet1.cell(row=1, column=i)
                cell.value = cell1.value
            us1 = create_user_list(sheet1)
            us2 = create_user_list(sheet2)
            print(us1)
            k = 2
            i = 0
            while i < len(us1):
                if not (us1[i] in us2):
                    for t in range(1, sheet1.max_column + 1):
                        cell = sheet.cell(row=k, column=t)
                        cell1 = sheet1.cell(row=k, column=t)
                        cell.value = cell1.value
                    k += 1
                i += 1
            if values[2] != '':
                wb.save(values[2])
            else:
                path = Path(pathlib.Path.home(), 'Documents', 'Итоговая_таблица.xlsx')
                print(path)
                wb.save(path)

            # print(event, values) #debug
            if event in (None, 'Exit', 'Cancel'):
                break
    except Exception as err:
        print(err)
        error_layout = [
            [sg.Text('Ошибка, проверьте файлы и заполненность форм ')
             ],
            [sg.Submit()]
        ]
        window.Close()
        window2 = sg.Window('File Compare', error_layout)
        while True:
            event = window2.read()
            if event in (None, 'Exit'):
                break
            else:
                window2.Close()
                main()
                break


if __name__ == '__main__':
    main()
